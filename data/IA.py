import csv
import random
import string
from openpyxl import Workbook
import os

# =========================================================
# CONFIG
# =========================================================

NB_EMPLOYES = 5000
NB_LOGS = 3000
DOSSIER = "game_data"

# Création du dossier principal
os.makedirs(DOSSIER, exist_ok=True)

# =========================================================
# DONNÉES
# =========================================================

noms = [
    "Martin", "Bernard", "Dubois", "Thomas",
    "Robert", "Richard", "Petit", "Durand",
    "Moreau", "Simon", "Laurent", "Michel"
]

prenoms = [
    "Lucas", "Emma", "Nathan", "Jade",
    "Hugo", "Lina", "Louis", "Sarah"
]

statuts = [
    "ACTIVE",
    "LOCKED",
    "ERROR",
    "DISCONNECTED",
    "BANNED"
]

actions_logs = [
    "LOGIN SUCCESS",
    "LOGIN FAILED",
    "ACCESS DENIED",
    "FILE OPENED",
    "DATABASE ERROR",
    "ADMIN CONNECTED",
    "ROOT LOGIN",
    "UNAUTHORIZED ACCESS",
    "PACKET LOST"
]

# =========================================================
# FONCTIONS
# =========================================================

def mot_de_passe(longueur=12):
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(longueur))


def cesar(message, decalage):
    resultat = ""

    for lettre in message:

        if lettre.isalpha():

            base = ord('A') if lettre.isupper() else ord('a')

            resultat += chr(
                (ord(lettre) - base + decalage) % 26 + base
            )

        else:
            resultat += lettre

    return resultat


# =========================================================
# CODES SECRETS
# =========================================================

code_secret_1 = "ALPHA-7"
code_secret_2 = "VX-11"
message_secret = "THE PASSWORD IS ORION"

message_code = cesar(message_secret, 3)

# =========================================================
# CSV EMPLOYÉS
# =========================================================

csv_path = os.path.join(DOSSIER, "employees.csv")

with open(csv_path, "w", newline="", encoding="utf-8") as f:

    writer = csv.writer(f)

    writer.writerow([
        "id",
        "prenom",
        "nom",
        "status",
        "password",
        "key"
    ])

    for i in range(NB_EMPLOYES):

        prenom = random.choice(prenoms)
        nom = random.choice(noms)
        status = random.choice(statuts)
        password = mot_de_passe()

        key = "NONE"

        # Indices cachés
        if i == 3487:
            key = code_secret_1

        if i == 4210:
            key = code_secret_2

        writer.writerow([
            i,
            prenom,
            nom,
            status,
            password,
            key
        ])

print("employees.csv généré")

# =========================================================
# LOGS TXT
# =========================================================

logs_path = os.path.join(DOSSIER, "server_logs.txt")

with open(logs_path, "w", encoding="utf-8") as f:

    for i in range(NB_LOGS):

        ligne = random.choice(actions_logs)

        # Faux timestamp
        heure = random.randint(0, 23)
        minute = random.randint(0, 59)
        seconde = random.randint(0, 59)

        log = f"[{heure:02}:{minute:02}:{seconde:02}] {ligne}"

        # Indices cachés
        if i == 1337:
            log = f"[12:51:03] OVERRIDE KEY = {code_secret_1}"

        if i == 2222:
            log = f"[03:14:08] ENCRYPTED MESSAGE = {message_code}"

        f.write(log + "\n")

print("server_logs.txt généré")

# =========================================================
# NOTES TXT
# =========================================================

notes_path = os.path.join(DOSSIER, "notes.txt")

notes = [
    "Ne pas oublier de changer le mot de passe admin.",
    "Le serveur backup plante parfois.",
    "Supprimer les anciens logs.",
    "Faire attention aux accès root.",
    "Projet ORION toujours actif.",
    "Le code VX-11 ouvre quelque chose.",
    "ALPHA-7 semble être une clé override."
]

with open(notes_path, "w", encoding="utf-8") as f:

    for _ in range(300):

        f.write(random.choice(notes) + "\n")

print("notes.txt généré")

# =========================================================
# EXCEL XLSX
# =========================================================

wb = Workbook()

# Feuille principale
ws = wb.active
ws.title = "Users"

ws.append([
    "ID",
    "USERNAME",
    "ACCESS_LEVEL",
    "TOKEN"
])

for i in range(2000):

    username = random.choice(prenoms).lower() + str(i)

    access = random.choice([
        "USER",
        "MODERATOR",
        "ADMIN"
    ])

    token = mot_de_passe(16)

    ws.append([
        i,
        username,
        access,
        token
    ])

# Indice caché
ws["D542"] = "OMEGA-ACCESS"

# Deuxième feuille
ws2 = wb.create_sheet(title="HiddenData")

ws2.append([
    "KEY",
    "VALUE"
])

for i in range(500):

    ws2.append([
        mot_de_passe(6),
        mot_de_passe(20)
    ])

# Indice caché
ws2["B201"] = "FINAL PASSWORD = ORION"

excel_path = os.path.join(DOSSIER, "admin_database.xlsx")

wb.save(excel_path)

print("admin_database.xlsx généré")

# =========================================================
# FICHIERS LEURRES
# =========================================================

for i in range(50):

    fake_file = os.path.join(
        DOSSIER,
        f"temp_{i}.txt"
    )

    with open(fake_file, "w", encoding="utf-8") as f:

        for _ in range(100):

            texte = mot_de_passe(40)

            f.write(texte + "\n")

print("Fichiers leurres générés")

# =========================================================
# README SECRET
# =========================================================

readme_path = os.path.join(DOSSIER, "README_IMPORTANT.txt")

with open(readme_path, "w", encoding="utf-8") as f:

    f.write("SYSTEM ACCESS PROJECT\n")
    f.write("---------------------\n\n")
    f.write("Some files contain hidden information.\n")
    f.write("Not all data is useful.\n")
    f.write("Search carefully.\n\n")
    f.write("Encrypted message:\n")
    f.write(message_code + "\n")

print("README_IMPORTANT.txt généré")

# =========================================================
# FIN
# =========================================================

print("\n===================================")
print("TOUS LES FICHIERS ONT ÉTÉ GÉNÉRÉS")
print("Dossier :", DOSSIER)
print("===================================")
